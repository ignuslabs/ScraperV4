"""Data export utilities for ScraperV4."""

import json
import csv
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
from datetime import datetime
import pandas as pd
from src.core.config import config

def export_to_json(data: Union[Dict[str, Any], List[Dict[str, Any]]], 
                  filename: Optional[str] = None, pretty: bool = True) -> str:
    """Export data to JSON format."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.json"
    
    export_path = Path(config.data_folder) / "exports" / filename
    export_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(export_path, 'w', encoding='utf-8') as f:
        if pretty:
            json.dump(data, f, indent=2, ensure_ascii=False)
        else:
            json.dump(data, f, ensure_ascii=False)
    
    return str(export_path)

def export_to_csv(data: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """Export data to CSV format (stub implementation)."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.csv"
    
    export_path = Path(config.data_folder) / "exports" / filename
    export_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not data:
        # Create empty CSV file
        with open(export_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['No data to export'])
        return str(export_path)
    
    # Get all unique keys from all records
    all_keys = set()
    for item in data:
        if isinstance(item, dict):
            all_keys.update(item.keys())
    
    # Write CSV
    with open(export_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=sorted(all_keys))
        writer.writeheader()
        
        for item in data:
            if isinstance(item, dict):
                # Handle nested data by converting to string
                flattened_item = {}
                for key, value in item.items():
                    if isinstance(value, (dict, list)):
                        flattened_item[key] = json.dumps(value, ensure_ascii=False)
                    else:
                        flattened_item[key] = value
                writer.writerow(flattened_item)
    
    return str(export_path)

def export_to_excel(data: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """Export data to Excel format using pandas and openpyxl."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.xlsx"
    
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename = filename.rsplit('.', 1)[0] + '.xlsx'
    
    export_path = Path(config.data_folder) / "exports" / filename
    export_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Handle empty data case
    if not data:
        # Create empty Excel file with header indicating no data
        empty_df = pd.DataFrame({"No Data": ["No records to export"]})
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            empty_df.to_excel(writer, sheet_name='Data', index=False)
        return str(export_path)
    
    try:
        # Flatten nested data structures for better Excel compatibility
        flattened_data = []
        for item in data:
            if isinstance(item, dict):
                flattened_item = flatten_nested_data(item)
                flattened_data.append(flattened_item)
            else:
                # Handle non-dict items by converting to string
                flattened_data.append({"value": str(item)})
        
        # Create DataFrame from flattened data
        df = pd.DataFrame(flattened_data)
        
        # Clean column names (remove special characters, handle duplicates)
        df.columns = [_clean_column_name(col) for col in df.columns]
        
        # Handle potential duplicate column names
        df.columns = _ensure_unique_column_names(df.columns)
        
        # Convert problematic data types for Excel compatibility
        df = _prepare_dataframe_for_excel(df)
        
        # Write to Excel using openpyxl engine
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Get the workbook and worksheet to apply formatting
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Apply basic formatting
            _format_excel_worksheet(worksheet, df)
        
        return str(export_path)
        
    except Exception as e:
        # If Excel export fails, fall back to JSON with error info
        fallback_path = export_path.with_suffix('.json')
        with open(fallback_path, 'w', encoding='utf-8') as f:
            json.dump({
                "error": f"Excel export failed: {str(e)}",
                "note": "Data exported as JSON fallback",
                "data": data
            }, f, indent=2, ensure_ascii=False)
        return str(fallback_path)

def clean_data_for_export(data: Union[Dict[str, Any], List[Dict[str, Any]]]) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
    """Clean data for export by removing None values and handling special types."""
    if isinstance(data, list):
        cleaned_list = []
        for item in data:
            if item is not None:
                cleaned_item = clean_data_for_export(item)
                # Only add dicts to the list, skip nested lists
                if isinstance(cleaned_item, dict):
                    cleaned_list.append(cleaned_item)
                elif isinstance(cleaned_item, list):
                    # Flatten nested lists
                    cleaned_list.extend(cleaned_item)
        return cleaned_list
    elif isinstance(data, dict):
        cleaned = {}
        for key, value in data.items():
            if value is not None:
                if isinstance(value, (dict, list)):
                    cleaned_value = clean_data_for_export(value)
                    if cleaned_value:  # Only add if not empty
                        cleaned[key] = cleaned_value
                else:
                    cleaned[key] = value
        return cleaned
    else:
        return data

def flatten_nested_data(data: Dict[str, Any], parent_key: str = '', separator: str = '_') -> Dict[str, Any]:
    """Flatten nested dictionary structure for easier CSV export."""
    items = []
    
    for key, value in data.items():
        new_key = f"{parent_key}{separator}{key}" if parent_key else key
        
        if isinstance(value, dict):
            items.extend(flatten_nested_data(value, new_key, separator).items())
        elif isinstance(value, list):
            # Handle lists by converting to string or creating indexed fields
            if value and isinstance(value[0], dict):
                for i, item in enumerate(value):
                    items.extend(flatten_nested_data(item, f"{new_key}_{i}", separator).items())
            else:
                items.append((new_key, json.dumps(value, ensure_ascii=False)))
        else:
            items.append((new_key, value))
    
    return dict(items)

def get_export_statistics(export_path: str) -> Dict[str, Any]:
    """Get statistics about an exported file."""
    path = Path(export_path)
    
    if not path.exists():
        return {"error": "Export file not found"}
    
    stats = {
        "filename": path.name,
        "file_size": path.stat().st_size,
        "created_at": datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
        "format": path.suffix.lower(),
        "path": str(path)
    }
    
    # Try to get record count for different file types
    if path.suffix.lower() == '.json':
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    stats["record_count"] = len(data)
                elif isinstance(data, dict):
                    stats["record_count"] = 1
        except Exception:
            stats["record_count"] = "unknown"
    elif path.suffix.lower() == '.xlsx':
        try:
            df = pd.read_excel(path)
            stats["record_count"] = len(df)
        except Exception:
            stats["record_count"] = "unknown"
    elif path.suffix.lower() == '.csv':
        try:
            df = pd.read_csv(path)
            stats["record_count"] = len(df)
        except Exception:
            stats["record_count"] = "unknown"
    
    return stats

def list_exports() -> List[Dict[str, Any]]:
    """List all export files with their statistics."""
    exports_dir = Path(config.data_folder) / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    
    exports = []
    for export_file in exports_dir.glob("*"):
        if export_file.is_file():
            stats = get_export_statistics(str(export_file))
            exports.append(stats)
    
    return sorted(exports, key=lambda x: x.get('created_at', ''), reverse=True)

def cleanup_old_exports(days_old: int = 30) -> int:
    """Clean up export files older than specified days."""
    from datetime import timedelta
    
    exports_dir = Path(config.data_folder) / "exports"
    if not exports_dir.exists():
        return 0
    
    cutoff_date = datetime.now() - timedelta(days=days_old)
    deleted_count = 0
    
    for export_file in exports_dir.glob("*"):
        if export_file.is_file():
            created_time = datetime.fromtimestamp(export_file.stat().st_ctime)
            if created_time < cutoff_date:
                export_file.unlink()
                deleted_count += 1
    
    return deleted_count


def _clean_column_name(column_name: str) -> str:
    """Clean column name for Excel compatibility."""
    import re
    
    # Convert to string if not already
    column_name = str(column_name)
    
    # Replace problematic characters with underscores
    cleaned = re.sub(r'[^\w\s-]', '_', column_name)
    
    # Replace multiple spaces/underscores with single underscore
    cleaned = re.sub(r'[\s_]+', '_', cleaned)
    
    # Remove leading/trailing underscores
    cleaned = cleaned.strip('_')
    
    # Ensure column name is not empty
    if not cleaned:
        cleaned = "column"
    
    # Truncate if too long (Excel has 255 char limit for column names)
    if len(cleaned) > 255:
        cleaned = cleaned[:252] + "..."
    
    return cleaned


def _ensure_unique_column_names(columns) -> List[str]:
    """Ensure all column names are unique by appending numbers to duplicates."""
    unique_columns = []
    column_counts = {}
    
    for col in columns:
        if col in column_counts:
            column_counts[col] += 1
            unique_col = f"{col}_{column_counts[col]}"
        else:
            column_counts[col] = 0
            unique_col = col
        
        unique_columns.append(unique_col)
    
    return unique_columns


def _prepare_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare DataFrame for Excel export by handling problematic data types."""
    df_copy = df.copy()
    
    for column in df_copy.columns:
        # Handle None/NaN values
        df_copy[column] = df_copy[column].fillna('')
        
        # Convert complex data types to strings
        df_copy[column] = df_copy[column].apply(lambda x: _convert_value_for_excel(x))
    
    return df_copy


def _convert_value_for_excel(value: Any) -> Any:
    """Convert individual values for Excel compatibility."""
    # Handle None
    if value is None:
        return ""
    
    # Handle complex data types by converting to JSON string
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    
    # Handle datetime objects
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    
    # Handle very long strings (Excel has cell content limits)
    if isinstance(value, str) and len(value) > 32767:  # Excel's cell character limit
        return value[:32764] + "..."
    
    # Return as-is for simple types
    return value


def _format_excel_worksheet(worksheet, df: pd.DataFrame) -> None:
    """Apply basic formatting to Excel worksheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Format header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Apply header formatting
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            # Calculate reasonable column width based on content
            max_length = max(
                len(str(column)),  # Header length
                max(len(str(val)) for val in df[column][:100] if val is not None) if len(df) > 0 else 0  # Sample data length
            )
            
            # Set reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except ImportError:
        # If openpyxl styles are not available, skip formatting
        pass
    except Exception:
        # If any formatting fails, continue without formatting
        pass